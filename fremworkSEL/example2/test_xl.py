import openpyxl
# give the xl file path
book = openpyxl.load_workbook("/home/sachin/Documents/demoPython.xlsx")
sheet = book.active
# for display the text
cell = sheet.cell(row= 1 , column= 2)
print(cell.value)
dict  = {}
# for insert the text
sheet.cell(row=2 , column=2).value="sachin"
print(sheet.cell(row=2 , column=2).value)
# print the number of column / row
print(sheet.max_row)
print(sheet.max_column)


for i in range(1 ,sheet.max_row+1):
    # if sheet.cell(row=i ,column=1 ).value == "CASE2":
        for j in range(1 ,sheet.max_column):
            # print(sheet.cell(row = i , column=j).value)

            dict[sheet.cell(row = 1 , column=j).value] =sheet.cell(row = i , column=j).value

            print(dict)
    
        